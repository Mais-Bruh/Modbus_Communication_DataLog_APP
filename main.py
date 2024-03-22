import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from pyModbusTCP.client import ModbusClient
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Function to handle button click event
def start_polling():
    # Get IP address, start address, register count, polling interval, and Excel file from user input
    plc_ip = ip_entry.get()
    start_address = int(start_address_entry.get())
    register_count = int(register_count_entry.get())
    polling_interval = polling_interval_var.get()
    excel_file = excel_file_entry.get()

    # Create a Modbus TCP client instance
    client = ModbusClient(host=plc_ip, port=502)

    # Initialize poll count
    poll_count = 0

    while True:
        if client.open():
            print("Connected to PLC")

            # Read input registers
            result = client.read_input_registers(start_address, register_count)
            if result:
                # Load the workbook and select the active worksheet
                wb = load_workbook(filename=excel_file)
                ws = wb.active

                # Determine the next available row (avoid overwriting existing data)
                next_row = ws.max_row + 1

                # Get current date and time
                now = datetime.now()
                current_date = now.date()
                current_time = now.strftime("%H:%M:%S")

                # Write date and time to the first and second columns, respectively
                ws.cell(row=next_row, column=1, value=str(current_date))
                ws.cell(row=next_row, column=2, value=str(current_time))

                # Append register data starting from the third column
                for col, value in enumerate(result, start=3):
                    ws.cell(row=next_row, column=col, value=value)

                # Save the workbook
                wb.save(excel_file)
                poll_count += 1
                print(f"Data saved to {excel_file} - Poll #: {poll_count}")

            else:
                print("Failed to read input registers")

            # Ensure to close the client
            client.close()

        else:
            print("Failed to connect to PLC")

        # Wait for the selected polling interval before the next poll
        time.sleep(polling_interval)

# Function to handle button click event to select Excel file
def select_excel_file():
    excel_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(0, excel_file)

# Create a Tkinter window
root = tk.Tk()
root.title("Modbus Polling GUI")

# Create IP address label and entry
ip_label = ttk.Label(root, text="PLC IP Address:")
ip_label.grid(row=0, column=0, padx=5, pady=5)
ip_entry = ttk.Entry(root)
ip_entry.grid(row=0, column=1, padx=5, pady=5)

# Create start address label and entry
start_address_label = ttk.Label(root, text="Start Address:")
start_address_label.grid(row=1, column=0, padx=5, pady=5)
start_address_entry = ttk.Entry(root)
start_address_entry.grid(row=1, column=1, padx=5, pady=5)

# Create register count label and entry
register_count_label = ttk.Label(root, text="Register Count:")
register_count_label.grid(row=2, column=0, padx=5, pady=5)
register_count_entry = ttk.Entry(root)
register_count_entry.grid(row=2, column=1, padx=5, pady=5)

# Create polling interval label and dropdown menu
polling_interval_label = ttk.Label(root, text="Polling Interval:")
polling_interval_label.grid(row=3, column=0, padx=5, pady=5)
polling_interval_var = tk.IntVar()
polling_interval_options = [30, 60, 120, 300]  # Polling intervals in seconds
for i, interval in enumerate(polling_interval_options):
    ttk.Radiobutton(root, text=f"{interval} sec", variable=polling_interval_var, value=interval).grid(row=3, column=i+1, padx=5, pady=5)

# Create Excel file label, entry, and button
excel_file_label = ttk.Label(root, text="Excel File:")
excel_file_label.grid(row=4, column=0, padx=5, pady=5)
excel_file_entry = ttk.Entry(root)
excel_file_entry.grid(row=4, column=1, padx=5, pady=5)
excel_file_button = ttk.Button(root, text="Browse", command=select_excel_file)
excel_file_button.grid(row=4, column=2, padx=5, pady=5)

# Create start button
start_button = ttk.Button(root, text="RUN", command=start_polling)
start_button.grid(row=5, column=0, columnspan=3, padx=5, pady=5)

# Run the Tkinter event loop
root.mainloop()

